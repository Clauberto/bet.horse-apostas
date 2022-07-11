from tkinter import *
import openpyxl

class BetHorseAdminLogged():

    book = openpyxl.Workbook()
    print(book.sheetnames)
    book.create_sheet('Cavalos')
    cavalos_page = book['Cavalos']


    def __init__(self, logged):
        self.logged = logged
        self.logged.geometry("500x500")
        self.logged.title("bet.horse: Administração")
        self.__nCavalo = StringVar()
        self.chance = DoubleVar()
        self.debug()

    @property
    def nCavalo(self):
        return self.__nCavalo

    @nCavalo.setter
    def nCavalo(self,novo_nCavalo):
        raise ValueError("Impossível alterar o nCavalo diretamente, use o metodo!")

    def debug(self):
        Label(self.logged, text="Número do Cavalo").place(x=175, y=250)
        __nCavalo = Entry(self.logged, textvariable=self.__nCavalo)
        __nCavalo.place(x=175, y=270)


        Label(self.logged, text="Chances de Vencer").place(x=175, y=300)
        chance = Entry(self.logged, textvariable=self.chance)
        chance.place(x=175, y=320)

        
        Button(self.logged, text="Registrar", command=self.setar, bg="red", fg="white", width=15).place(x=175, y=350)
        
    def setar(self):
        nCavalo = self.__nCavalo.get()
        chance = self.chance.get()
        print(self.__nCavalo.get(), self.chance.get())
        self.cavalos_page.cell(row=1, column=1).value = 'Cavalos'
        self.cavalos_page.cell(row=1, column=2).value = 'Chances'
        self.cavalos_page.append([nCavalo, chance])
        self.book.save('BetHorse.xlsx')
